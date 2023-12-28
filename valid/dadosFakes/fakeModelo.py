import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import random

# Criar um novo arquivo XLSX
wb = openpyxl.Workbook()

# Selecionar a primeira planilha
sheet = wb.active

# Definir cabeçalho
header = ['A', 'B', 'C', 'D','E']
for col_num, col_letter in enumerate(header, 1):
    sheet[get_column_letter(col_num) + '1'] = col_letter
    sheet[get_column_letter(col_num) + '1'].font = Font(bold=True)

# Preencher as linhas de 2 a 11 com números aleatórios
for row_num in range(2, 14):
    for col_num in range(1, 6):
        sheet[get_column_letter(col_num) + str(row_num)] = random.randint(1, 100)

# Salvar o arquivo
wb.save('arquivo_excel.xlsx')
